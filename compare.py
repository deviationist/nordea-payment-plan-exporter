#!/usr/bin/env python3
"""Compare the INITIAL payment plan against the REDUCED (post-downpayment) plan and
show, month by month: the bank's two payments, your monthly top-up (you keep
paying the initial amount), and your forsprang vs the initial plan.

Two sources for the reduced plan:
  - estimate (default): re-annuitise (initial balance - downpayment) over the term
  - real:  --post <timestamp>  uses Nordea's captured post-downpayment export

The forsprang is always measured against the INITIAL plan: it's how far below the
initial plan's balance you run by keeping the initial payment on the reduced loan.

    ./compare.sh                       # estimate, committed baseline as initial
    ./compare.sh --post <ts>           # hard comparison vs a captured reduced plan
    ./compare.sh --downpayment 500000 --format xlsx

NOTE: the annuity helpers mirror build_sheet.py (kept self-contained on purpose so
the interactive model and this report stay independent).
"""
import argparse
import csv as csvmod
import json
from datetime import date
from pathlib import Path

import envcfg

HERE = Path(__file__).resolve().parent
XLSX_OUT = "nordea-sammenligning.xlsx"
CSV_OUT = "nordea-sammenligning.csv"


def pmt(r, n, pv):
    return pv * r / (1 - (1 + r) ** -n) if r else pv / n


def add_months(d, k):
    m = d.month - 1 + k
    return date(d.year + m // 12, m % 12 + 1, d.day)


def _src(ts, stem_detail, stem_pay):
    if ts:
        return f"captures/loan-detail-{ts}.json", f"captures/payplan-{ts}.json"
    return stem_detail, stem_pay


def load_plan(ts):
    """Return the annuity parameters of a captured/baseline plan."""
    dp, pp = _src(ts, "nordea_loan_detail.json", "nordea_payplan_data.json")
    d = json.loads((HERE / dp).read_text())
    bank = json.loads((HERE / pp).read_text())
    rs = d["repayment_schedule"]
    principal = d["amount"]["granted"]
    rate = d["interest"]["base_rate"] / 100
    tpy = rs["terms_per_year"]
    first_due = date.fromisoformat(rs["following_instalment"])
    fee = d["following_payment"]["fees"]
    terms = rs["number_of_instalments"]
    factor = bank["pay_plans"][0]["interest"] / (principal * rate / tpy)
    return dict(principal=principal, rate=rate, tpy=tpy, first_due=first_due,
                fee=fee, terms=terms, factor=factor,
                balance=abs(d["amount"]["balance"]))


def months_between(a, b):
    return (b.year - a.year) * 12 + (b.month - a.month)


def schedule(opening, annual, n_terms, first_due, fee, factor, tpy=12, fixed=None, adj=None):
    """Monthly rows until payoff. fixed=None re-annuitises (the plan's own payment);
    fixed=<amount ex-fee> pays a constant amount (your-actual). factor scales term-1
    interest (partial first period). adj = {term_index: signed_delta} applies a
    balance change at that term's start — the window: -amount at the downpayment,
    +amount at the uploan."""
    adj = adj or {}
    rows, bal = [], opening
    for i in range(n_terms):
        bal += adj.get(i, 0.0)
        if bal <= 0.005:
            break
        mr = annual / tpy
        full_int = bal * mr
        billed_int = full_int * (factor if i == 0 else 1.0)   # partial first period
        if fixed is None:
            prin = pmt(mr, n_terms - i, bal) - full_int       # normal annuity principal
        else:
            prin = fixed - billed_int                         # your-actual: fixed payment
        prin = min(prin, bal)                                 # final (partial) term
        bal -= prin
        rows.append(dict(date=add_months(first_due, i), total=prin + billed_int + fee,
                         interest=billed_int, principal=prin, balance=max(bal, 0.0)))
    return rows


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--capture", metavar="TS", help="initial plan source (default: baseline)")
    ap.add_argument("--post", metavar="TS", help="captured reduced plan (hard comparison)")
    ap.add_argument("--downpayment", type=float, default=envcfg.get_float("DOWNPAYMENT", 0.0),
                    metavar="KR",
                    help="extra downpayment for the estimate (default: .env DOWNPAYMENT, else 0)")
    ap.add_argument("--down-date", type=date.fromisoformat, metavar="YYYY-MM-DD",
                    help="when the downpayment was made (default: first term)")
    ap.add_argument("--up-date", type=date.fromisoformat, metavar="YYYY-MM-DD",
                    help="when the uploan (opplåning) returns it (default: 4 yr after down)")
    ap.add_argument("--format", choices=["both", "csv", "xlsx"], default="both")
    args = ap.parse_args()

    A = load_plan(args.capture)
    # the amount you keep paying (initial plan's steady annuity, ex-fee)
    keep_paying = pmt(A["rate"] / A["tpy"], A["terms"], A["principal"])

    down_date = args.down_date or envcfg.get_date("DOWN_DATE", A["first_due"])
    up_date = args.up_date or envcfg.get_date("UP_DATE", add_months(down_date, 48))
    down_idx = months_between(A["first_due"], down_date)
    up_idx = months_between(A["first_due"], up_date)

    if args.post:                                   # hard comparison: real reduced plan
        B = load_plan(args.post)
        reduced_open, reduced_rate = B["balance"], B["rate"]
        reduced_terms, reduced_first = B["terms"], B["first_due"]
        reduced_factor, mode = B["factor"], f"bankens reduserte plan ({args.post})"
    else:                                           # estimate: re-annuitise minus downpayment
        reduced_open = A["principal"] - args.downpayment
        reduced_rate, reduced_terms = A["rate"], A["terms"]
        reduced_first, reduced_factor = A["first_due"], A["factor"]
        mode = f"estimat (−{args.downpayment:,.0f})"

    plan_a = schedule(A["principal"], A["rate"], A["terms"], A["first_due"],
                      A["fee"], A["factor"], A["tpy"])
    plan_b = schedule(reduced_open, reduced_rate, reduced_terms, reduced_first,
                      A["fee"], reduced_factor, A["tpy"])
    # your reality: keep paying the initial amount, but the downpayment is only deployed
    # DURING the window — in at the downpayment, back out at the uploan (opplåning)
    yours = schedule(A["principal"], A["rate"], A["terms"], A["first_due"],
                     A["fee"], A["factor"], A["tpy"], fixed=keep_paying,
                     adj={down_idx: -args.downpayment, up_idx: +args.downpayment})

    a_by = {r["date"]: r for r in plan_a}
    b_by = {r["date"]: r for r in plan_b}
    y_by = {r["date"]: r for r in yours}
    dates = sorted(set(a_by) | set(b_by))

    rows = []
    for d in dates:
        a, b, y = a_by.get(d), b_by.get(d), y_by.get(d)
        a_bal = a["balance"] if a else 0.0
        y_bal = y["balance"] if y else 0.0
        topup = (a["total"] if a else 0.0) - (b["total"] if b else 0.0)
        rows.append(dict(
            date=d,
            a_total=a["total"] if a else 0.0, a_bal=a_bal,
            b_total=b["total"] if b else 0.0, b_bal=b["balance"] if b else 0.0,
            topup=topup, your_bal=y_bal, forsprang=a_bal - y_bal))

    # summary
    def payoff(sched):
        return sched[-1]["date"] if sched else None
    up_term_date = add_months(A["first_due"], up_idx)
    forsprang_after = next((r["forsprang"] for r in rows if r["date"] == up_term_date), 0.0)
    summary = dict(
        mode=mode, down_date=down_date, up_date=up_date,
        a_payoff=payoff(plan_a), b_payoff=payoff(plan_b), your_payoff=payoff(yours),
        months_saved=len(plan_a) - len(yours),
        keep_paying=keep_paying + A["fee"],
        b_payment=plan_b[0]["total"] if plan_b else 0.0,
        max_forsprang=max((r["forsprang"] for r in rows), default=0.0),
        forsprang_after=forsprang_after,
    )
    # report the steady reduced payment (term 2), not the partial first-term stub
    summary["b_payment"] = plan_b[1]["total"] if len(plan_b) > 1 else summary["b_payment"]
    write_outputs(rows, summary, args.format)
    print(f"mode: {mode}")
    print(f"  window:          {down_date} → {up_date} (uploan)")
    print(f"  initial payoff:  {summary['a_payoff']}")
    print(f"  reduced payoff:  {summary['b_payoff']}  (bank payment {summary['b_payment']:,.0f})")
    print(f"  YOUR payoff:     {summary['your_payoff']}  (you pay {summary['keep_paying']:,.0f})")
    print(f"  months saved vs initial: {summary['months_saved']}")
    print(f"  forsprang during window (max): {summary['max_forsprang']:,.0f}")
    print(f"  lasting forsprang after uploan: {summary['forsprang_after']:,.0f}")


def write_outputs(rows, summary, fmt):
    def no(v):
        return f"{v:.2f}".replace(".", ",")
    if fmt in ("both", "csv"):
        with open(HERE / CSV_OUT, "w", encoding="utf-8-sig", newline="") as fh:
            w = csvmod.writer(fh, delimiter=";")
            w.writerow(["Dato", "Initial terminbeløp", "Initial restgjeld",
                        "Redusert terminbeløp", "Redusert restgjeld",
                        "Din topup (initial−redusert)", "Din restgjeld (betaler initial)",
                        "Forsprang vs initial"])
            for r in rows:
                w.writerow([r["date"].strftime("%d.%m.%Y"), no(r["a_total"]), no(r["a_bal"]),
                            no(r["b_total"]), no(r["b_bal"]), no(r["topup"]),
                            no(r["your_bal"]), no(r["forsprang"])])
        print(f"Wrote {CSV_OUT} · {len(rows)} rows")

    if fmt in ("both", "xlsx"):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Sammenligning"
        ws["A1"] = "Initial vs redusert nedbetalingsplan"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Redusert plan: {summary['mode']}"
        ws["A2"].font = Font(italic=True, color="666666")
        info = [("Nedbetaling (downpayment)", summary["down_date"]),
                ("Opplåning (uploan)", summary["up_date"]),
                ("Initial innfridd", summary["a_payoff"]),
                ("Redusert innfridd", summary["b_payoff"]),
                ("Din innfrielse (betaler initial)", summary["your_payoff"]),
                ("Måneder spart vs initial", summary["months_saved"]),
                ("Du betaler / mnd", summary["keep_paying"]),
                ("Bankens reduserte / mnd", summary["b_payment"]),
                ("Maks forsprang (i vinduet)", summary["max_forsprang"]),
                ("Varig forsprang etter opplåning", summary["forsprang_after"])]
        for i, (k, v) in enumerate(info, start=4):
            ws.cell(row=i, column=1, value=k).font = Font(bold=True)
            c = ws.cell(row=i, column=2, value=v)
            if isinstance(v, float):
                c.number_format = "#,##0.00"
            elif isinstance(v, date):
                c.number_format = "dd.mm.yyyy"
        hr = 4 + len(info) + 1
        headers = ["Dato", "Initial\nterminbeløp", "Initial\nrestgjeld",
                   "Redusert\nterminbeløp", "Redusert\nrestgjeld",
                   "Topup\n(init−red)", "Din restgjeld\n(betaler initial)",
                   "Forsprang\nvs initial"]
        blue, white = PatternFill("solid", fgColor="00427A"), Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=hr, column=col, value=h)
            c.fill, c.font = blue, white
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in rows:
            ws.append([r["date"], r["a_total"], r["a_bal"], r["b_total"], r["b_bal"],
                       r["topup"], r["your_bal"], r["forsprang"]])
        for row in ws.iter_rows(min_row=hr + 1, min_col=1, max_col=8):
            row[0].number_format = "dd.mm.yyyy"
            for c in row[1:]:
                c.number_format = "#,##0.00"
        for i, wdt in enumerate([13, 14, 15, 14, 15, 13, 15, 14], start=1):
            ws.column_dimensions[get_column_letter(i)].width = wdt
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)
        wb.save(HERE / XLSX_OUT)
        print(f"Wrote {XLSX_OUT} · {len(rows)} rows")


if __name__ == "__main__":
    main()
