#!/usr/bin/env python3
"""Build a dynamic, formula-driven Nordea mortgage workbook.

Sheets:
  1. "Modell (dynamisk)"      – inputs + a full month-by-month annuity schedule
                                 driven entirely by Excel formulas. Each month has
                                 its own editable nominal-rate cell that carries
                                 forward; change one and everything below recomputes.
  2. "Lånedetaljer"          – static loan facts pulled from the bank.
  3. "Bankens plan (original)" – the bank's own 47-row published plan, as extracted
                                 (static reference; computed at 4,99 %).
"""
import argparse
import csv as csvmod
import json
import math
import os
import re
import zipfile
from datetime import date

import envcfg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

XLSX_OUT = "nordea-nedbetalingsplan.xlsx"
CSV_OUT = "nordea-nedbetalingsplan.csv"

_ap = argparse.ArgumentParser(
    description="Build the Nordea payment-plan CSV and/or XLSX from the raw JSON sources.")
_ap.add_argument("--format", choices=["both", "csv", "xlsx"], default="both",
                 help="which output(s) to produce (default: both)")
_ap.add_argument("--capture", metavar="TIMESTAMP",
                 help="build from a captures/<...>-TIMESTAMP.json snapshot "
                      "(default: the committed baseline nordea_*.json)")
_ap.add_argument("--downpayment", type=float, default=envcfg.get_float("DOWNPAYMENT", 0.0),
                 metavar="KR",
                 help="temporary extra-downpayment for the estimate scenario "
                      "(default: .env DOWNPAYMENT, else 0 = initial plan only)")
_args = _ap.parse_args()
WANT_CSV = _args.format in ("both", "csv")
WANT_XLSX = _args.format in ("both", "xlsx")

# ---- raw bank payloads ----
#   loan detail -> GET /api/dbf/ca/loans-v1/loans/<id>
#   pay-plans   -> GET /api/dbf/ca/loans-v1/loans/<id>/pay-plans
# Default: the committed baseline files. With --capture <ts>: a gitignored,
# timestamped snapshot written by fetch.py (refresh figures without re-login).
if _args.capture:
    LOAN_DETAIL_SRC = f"captures/loan-detail-{_args.capture}.json"
    PAYPLAN_SRC = f"captures/payplan-{_args.capture}.json"
else:
    LOAN_DETAIL_SRC = "nordea_loan_detail.json"
    PAYPLAN_SRC = "nordea_payplan_data.json"

with open(LOAN_DETAIL_SRC, encoding="utf-8") as fh:
    _d = json.load(fh)
_rs = _d["repayment_schedule"]

# loan facts derived from the raw detail payload (single source of truth)
LOAN = {
    "loan_id": _d["loan_id"],
    "formatted_id": _d["loan_formatted_id"],
    "iban": _d["iban"],
    "product_code": _d["product_code"],
    "nickname": _d["nickname"],
    "principal": _d["amount"]["granted"],
    "nominal_rate": _d["interest"]["base_rate"] / 100,   # applied to balance
    "effective_rate": _d["interest"]["rate"] / 100,      # effektiv (incl. fee)
    "fee_per_term": _d["following_payment"]["fees"],
    "terms_total": _rs["number_of_instalments"],
    "terms_per_year": _rs["terms_per_year"],
    "first_due": date.fromisoformat(_rs["following_instalment"]),
    "final_due": date.fromisoformat(_rs["final_payment_date"]),
    "debit_account": _rs["debit_account_number"],
    "owner": _d["owners"][0]["name"],
}

# ---- bank's published plan + first-term partial-period factor ----
with open(PAYPLAN_SRC, encoding="utf-8") as fh:
    BANK = json.load(fh)

# Nordea's first term (e.g. 15.07.2026) is a short partial period: interest is
# billed for only ~20 of 30 days because the loan was disbursed mid-month. The
# avdrag (principal) and resulting balance are already the normal full-annuity
# values, so only the FIRST term's interest needs scaling. Derive that fraction
# straight from the bank data = (billed first interest) / (a full month's interest).
_full_first_interest = LOAN["principal"] * LOAN["nominal_rate"] / LOAN["terms_per_year"]
FIRST_FACTOR = BANK["pay_plans"][0]["interest"] / _full_first_interest  # ~0.667 (20/30)

# ---- temporary extra-downpayment window (lump sum, paid in then redrawn) ----
WINDOW_AMOUNT = _args.downpayment   # paid in at the window start, redrawn at the end; 0 = none
WINDOW_MONTHS = 48            # default window length (~4 years)


def _addm(d, k):
    m = d.month - 1 + k
    return date(d.year + m // 12, m % 12 + 1, d.day)


# window dates: .env DOWN_DATE / UP_DATE, else first term and +4 years
WINDOW_START = envcfg.get_date("DOWN_DATE", LOAN["first_due"])
WINDOW_END = envcfg.get_date("UP_DATE", _addm(LOAN["first_due"], WINDOW_MONTHS))

# ---- styles ----
BLUE = PatternFill("solid", fgColor="00427A")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
ITALIC_GREY = Font(italic=True, color="666666")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "#,##0.00"
PCT = "0.00%"
DATEF = "dd.mm.yyyy"
INPUT_FILL = PatternFill("solid", fgColor="FFF6D9")  # pale yellow = editable

wb = Workbook()

# =====================================================================
# Sheet 1 — dynamic model
# =====================================================================
ws = wb.active
ws.title = "Modell (dynamisk)"

ws["A1"] = "Nedbetalingsplan – dynamisk modell"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = (f"Lån {LOAN['formatted_id']} «{LOAN['nickname']}» · annuitetslån · "
            f"kilde: Nordea Nettbank")
ws["A2"].font = ITALIC_GREY

# --- inputs block (col A label, col B value) ---
inputs = [
    (4,  "Opprinnelig lånebeløp (kr)", LOAN["principal"], MONEY, True),
    (5,  "Antall terminer (totalt)",   LOAN["terms_total"], "0", True),
    (6,  "Terminer per år",            LOAN["terms_per_year"], "0", True),
    (7,  "Gebyr per termin (kr)",      LOAN["fee_per_term"], MONEY, True),
    (8,  "Første forfallsdato",        LOAN["first_due"], DATEF, True),
    (9,  "Nominell årsrente (standard)", LOAN["nominal_rate"], PCT, True),
    (10, "Andel rente 1. termin (≈ dager/30)", FIRST_FACTOR, PCT, True),
]
for row, label, val, fmt, editable in inputs:
    ws.cell(row=row, column=1, value=label).font = BOLD
    c = ws.cell(row=row, column=2, value=val)
    c.number_format = fmt
    c.border = BORDER
    if editable:
        c.fill = INPUT_FILL

# computed outputs
outputs = [
    # effektiv rente rounded UP to 2 decimals (the bank's reporting convention)
    (11, "Effektiv rente (kun renters rente)", "=CEILING((1+B9/B6)^B6-1,0.0001)", PCT),
    (12, "Effektiv rente (inkl. gebyr, IRR)",  "=CEILING((1+IRR(K15:K375,0.004))^12-1,0.0001)", PCT),
    (13, "Sluttdato (siste termin)",           "=EDATE(B8,B5-1)", DATEF),
    (14, "Sum renter over hele løpet",         "=SUM(F16:F375)", MONEY),
]
for row, label, formula, fmt in outputs:
    ws.cell(row=row, column=1, value=label).font = ITALIC_GREY
    c = ws.cell(row=row, column=2, value=formula)
    c.number_format = fmt

# --- temporary extra-downpayment window (scenario inputs/helpers/summary, cols C/D) ---
ws["C4"] = "— Ekstra nedbetaling (vindu) —"
ws["C4"].font = BOLD
scenario_io = [
    (5,  "Beløp (kr)",            WINDOW_AMOUNT,  MONEY, True),
    (6,  "Startdato",             WINDOW_START,   DATEF, True),
    (7,  "Sluttdato",             WINDOW_END,     DATEF, True),
    (8,  "Starttermin (#)",       "=(YEAR($D$6)-YEAR($B$8))*12+MONTH($D$6)-MONTH($B$8)+1", "0", False),
    (9,  "Sluttermin (#)",        "=(YEAR($D$7)-YEAR($B$8))*12+MONTH($D$7)-MONTH($B$8)+1", "0", False),
    (11, "Innfridd (scenario)",   '=EDATE($B$8,COUNTIF(P16:P375,">1"))', DATEF, False),
    (12, "Måneder spart",         '=$B$5-COUNTIF(P16:P375,">1")-1',      "0", False),
    (13, "Spart rente totalt",    "=SUM(F16:F375)-SUM(N16:N375)",        MONEY, False),
    (14, "Forsprang etter vindu", "=INDEX(Q16:Q375,$D$9)",               MONEY, False),
]
for row, label, val, fmt, editable in scenario_io:
    lab = ws.cell(row=row, column=3, value=label)
    lab.font = BOLD if editable else ITALIC_GREY
    c = ws.cell(row=row, column=4, value=val)
    c.number_format = fmt
    c.border = BORDER
    if editable:
        c.fill = INPUT_FILL

# --- schedule ---
HEAD_ROW = 15
HEADERS = ["Termin", "Forfallsdato", "Nominell\nårsrente", "Inngående saldo",
           "Terminbeløp\n(ekskl. gebyr)", "Rente", "Avdrag", "Gebyr",
           "Totalt", "Restgjeld", "Kontantstrøm*"]
for col, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=HEAD_ROW, column=col, value=h)
    c.fill = BLUE
    c.font = WHITE_BOLD
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

# scenario column headers (L–Q), right of the hidden cashflow col K
SCEN_HEADERS = ["Justering\nvindu (±)", "Inngående saldo\n(scenario)", "Rente\n(scenario)",
                "Ekstra avdrag\npr. mnd", "Restgjeld\n(scenario)", "Forsprang\nvs opprinnelig"]
for off, h in enumerate(SCEN_HEADERS):
    c = ws.cell(row=HEAD_ROW, column=12 + off, value=h)
    c.fill = BLUE
    c.font = WHITE_BOLD
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

# cashflow time-0 (disbursement) sits on the header row in col K
ws.cell(row=HEAD_ROW, column=11, value="=-$B$4").number_format = MONEY

n = LOAN["terms_total"]
first = HEAD_ROW + 1            # 15
last = HEAD_ROW + n            # 374
for i in range(n):
    r = first + i
    p = r - 1                   # previous row
    A = f"A{r}"
    if i == 0:
        ws[f"A{r}"] = 1
        ws[f"B{r}"] = "=$B$8"
        ws[f"C{r}"] = "=$B$9"
        ws[f"D{r}"] = "=$B$4"
    else:
        ws[f"A{r}"] = f"=A{p}+1"
        ws[f"B{r}"] = f"=EDATE(B{p},1)"
        ws[f"C{r}"] = f"=C{p}"            # carry rate forward (override propagates)
        ws[f"D{r}"] = f"=J{p}"            # opening = previous closing balance
    ws[f"H{r}"] = "=$B$7"                 # fee
    if i == 0:
        # First term is a partial period: avdrag follows the normal full-month
        # annuity (so the balance lands where the bank's does), but interest is
        # only billed for the partial period (factor in $B$10).
        ws[f"G{r}"] = f"=-PMT(C{r}/$B$6,$B$5-A{r}+1,D{r})-D{r}*C{r}/$B$6"
        ws[f"F{r}"] = f"=D{r}*C{r}/$B$6*$B$10"   # partial interest billed
        ws[f"E{r}"] = f"=G{r}+F{r}"              # actual (reduced) terminbeløp
    else:
        # re-annuitise over the REMAINING term at this month's rate
        ws[f"E{r}"] = f"=-PMT(C{r}/$B$6,$B$5-A{r}+1,D{r})"
        ws[f"F{r}"] = f"=D{r}*C{r}/$B$6"  # full-month interest on opening balance
        ws[f"G{r}"] = f"=E{r}-F{r}"       # principal
    ws[f"I{r}"] = f"=E{r}+H{r}"           # total incl. fee
    ws[f"J{r}"] = f"=D{r}-G{r}"           # closing balance
    ws[f"K{r}"] = f"=I{r}"                # cashflow for IRR

    # --- scenario columns: temporary downpayment window ---
    # L injects -amount at the start term and +amount (redraw) at the end term.
    # Interest accrues on the post-adjustment balance (M+L); the borrower still
    # pays the original ex-fee terminbeløp E, so principal = E-N; balance floored
    # at 0 for early payoff. O = interest saved that month; Q = head start vs plan.
    ws[f"L{r}"] = f"=IF(A{r}=$D$8,-$D$5,0)+IF(A{r}=$D$9,$D$5,0)"
    ws[f"M{r}"] = "=$B$4" if i == 0 else f"=P{p}"
    if i == 0:
        ws[f"N{r}"] = f"=(M{r}+L{r})*C{r}/$B$6*$B$10"
        ws[f"P{r}"] = f"=MAX(0,(M{r}+L{r})-(E{r}-N{r}))"
    else:
        ws[f"N{r}"] = f"=IF((M{r}+L{r})<=0,0,(M{r}+L{r})*C{r}/$B$6)"
        ws[f"P{r}"] = f"=IF((M{r}+L{r})<=0,0,MAX(0,(M{r}+L{r})-(E{r}-N{r})))"
    ws[f"O{r}"] = f"=IF(M{r}<=0,0,F{r}-N{r})"
    ws[f"Q{r}"] = f"=J{r}-P{r}"

    # formats
    ws[f"A{r}"].number_format = "0"
    ws[f"A{r}"].alignment = Alignment(horizontal="center")
    ws[f"B{r}"].number_format = DATEF
    ws[f"C{r}"].number_format = PCT
    ws[f"C{r}"].fill = INPUT_FILL         # editable per-month rate
    for col in "DEFGHIJKLMNOPQ":
        ws[f"{col}{r}"].number_format = MONEY
    for col in range(1, 18):
        ws.cell(row=r, column=col).border = BORDER

# totals row
tot = last + 1
ws.cell(row=tot, column=1, value="Sum").font = BOLD
for col_letter, col_idx in (("F", 6), ("G", 7), ("H", 8), ("I", 9), ("N", 14)):
    c = ws.cell(row=tot, column=col_idx, value=f"=SUM({col_letter}{first}:{col_letter}{last})")
    c.font = BOLD
    c.number_format = MONEY
    c.border = BORDER

# legend
ws.cell(row=tot + 2, column=1,
        value=("Gule celler er redigerbare. Endre renten i en celle i kolonne C – "
               "alle påfølgende måneder arver den nye renten automatisk, og hele planen "
               "(terminbeløp, renter, restgjeld, sluttdato) regnes om.")).font = ITALIC_GREY
ws.cell(row=tot + 3, column=1,
        value="* Kontantstrøm-kolonnen brukes kun til å beregne effektiv rente (IRR).").font = ITALIC_GREY
ws.cell(row=tot + 4, column=1,
        value=("1. termin er en delperiode: avdraget følger normal annuitet, men renten "
               "belastes kun for andelen i celle B10 (≈ 20/30 dager), slik banken gjør.")).font = ITALIC_GREY

# widths / freeze / hide
widths = [8, 13, 11, 16, 15, 13, 13, 9, 14, 16, 14,   # A–K
          13, 16, 13, 14, 16, 16]                      # L–Q (scenario)
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.column_dimensions["K"].hidden = True
ws.freeze_panes = ws.cell(row=first, column=1)

# =====================================================================
# Sheet 2 — loan details (static)
# =====================================================================
wd = wb.create_sheet("Lånedetaljer")
wd["A1"] = "Lånedetaljer"
wd["A1"].font = Font(bold=True, size=14)
details = [
    ("Lån-ID", LOAN["loan_id"]),
    ("Kontonummer", LOAN["formatted_id"]),
    ("IBAN", LOAN["iban"]),
    ("Produktkode", LOAN["product_code"]),
    ("Type", "Boliglån · annuitetslån · flytende rente"),
    ("Kallenavn", LOAN["nickname"]),
    ("Nominell rente", LOAN["nominal_rate"]),
    ("Effektiv rente", LOAN["effective_rate"]),
    ("Innvilget beløp (kr)", LOAN["principal"]),
    ("Antall terminer", LOAN["terms_total"]),
    ("Terminer per år", LOAN["terms_per_year"]),
    ("Gebyr per termin (kr)", LOAN["fee_per_term"]),
    ("Første forfall", LOAN["first_due"]),
    ("Siste forfall", LOAN["final_due"]),
    ("Belastningskonto", LOAN["debit_account"]),
    ("Eier", LOAN["owner"]),
]
for i, (label, val) in enumerate(details, start=3):
    wd.cell(row=i, column=1, value=label).font = BOLD
    c = wd.cell(row=i, column=2, value=val)
    if isinstance(val, float):
        c.number_format = PCT if val < 1 else MONEY
    elif isinstance(val, date):
        c.number_format = DATEF
wd.column_dimensions["A"].width = 22
wd.column_dimensions["B"].width = 32
wd.cell(row=len(details) + 4, column=1,
        value=("NB: Bankens publiserte nedbetalingsplan er regnet med 4,99 % (nominell). "
               "Dagens effektive rente er 5,13 %. Bruk «Modell»-arket for å se reell plan "
               "ved gjeldende eller framtidig rente.")).font = ITALIC_GREY

# =====================================================================
# Sheet 3 — bank's original published plan (static reference)
# =====================================================================
with open("nordea_payplan_data.json", encoding="utf-8") as fh:
    bank = json.load(fh)

wb3 = wb.create_sheet("Bankens plan (original)")
wb3["A1"] = "Bankens publiserte nedbetalingsplan (uttrekk, statisk)"
wb3["A1"].font = Font(bold=True, size=12)
hdr = ["Forfallsdato", "Periode", "Totalt terminbeløp", "Avdrag", "Rente", "Gebyrer", "Restgjeld"]
HR = 3
for col, h in enumerate(hdr, start=1):
    c = wb3.cell(row=HR, column=col, value=h)
    c.fill = BLUE
    c.font = WHITE_BOLD
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

def classify(due):
    d = date.fromisoformat(due)
    return "År" if (d.month == 12 and d.day == 31) else "Måned"

r = HR
for p in bank["pay_plans"]:
    r += 1
    d = date.fromisoformat(p["due_date"])
    vals = [d, classify(p["due_date"]), p["total"], p["amount"],
            p["interest"], p["fee"], p["loan_balance"]]
    for col, v in enumerate(vals, start=1):
        c = wb3.cell(row=r, column=col, value=v)
        c.border = BORDER
        if col == 1:
            c.number_format = DATEF
        elif col == 2:
            c.alignment = Alignment(horizontal="center")
        else:
            c.number_format = MONEY
for i, w in enumerate([13, 9, 18, 14, 14, 11, 16], start=1):
    wb3.column_dimensions[get_column_letter(i)].width = w
wb3.freeze_panes = wb3.cell(row=HR + 1, column=1)

# =====================================================================
# Compute every cell value once. Used for (a) injecting cached results
# into the XLSX — openpyxl writes formulas with EMPTY <v/>, so viewers
# that don't recalculate (Numbers, Quick Look, Preview) would show 0 and
# bogus 31.12.1899 dates — and (b) writing the flat CSV export.
# =====================================================================
EPOCH = date(1899, 12, 30)


def _serial(d):
    return (d - EPOCH).days


def _add_months(d, k):
    m = d.month - 1 + k
    return date(d.year + m // 12, m % 12 + 1, d.day)


def _pmt(r, nper, pv):
    return pv * r / (1 - (1 + r) ** -nper) if r else pv / nper


def _ceil2(x):                          # round UP to 2 decimals (= Excel CEILING(x, 0.0001))
    return math.ceil(round(x * 10000, 4)) / 10000


rate = LOAN["nominal_rate"]
tpy = LOAN["terms_per_year"]
P = LOAN["principal"]
fee = LOAN["fee_per_term"]
first_due = LOAN["first_due"]

vals = {}
schedule_rows = []          # flat data for the CSV export
bal = P
cashflows = [-P]
sum_F = sum_G = sum_H = sum_I = 0.0

# scenario (temporary downpayment window) state — dates from config (.env)
win_start = WINDOW_START
win_end = WINDOW_END
start_term = (win_start.year - first_due.year) * 12 + (win_start.month - first_due.month) + 1
end_term = (win_end.year - first_due.year) * 12 + (win_end.month - first_due.month) + 1
scen_bal = P
scen_int = 0.0
scen_payoff_count = 0
forsprang_after = None

for i in range(n):
    r = first + i
    term = i + 1
    mr = rate / tpy
    rem = n - i
    annuity = _pmt(mr, rem, bal)
    full_int = bal * mr
    if i == 0:
        # partial first period: principal as normal annuity, interest scaled
        G = annuity - full_int
        F = full_int * FIRST_FACTOR
        E = G + F
    else:
        E = annuity
        F = full_int
        G = E - F
    J = bal - G
    Itot = E + fee
    if term > 1:
        vals[f"A{r}"] = term          # first term is a literal, skip
    vals[f"B{r}"] = _serial(_add_months(first_due, i))
    vals[f"C{r}"] = rate
    vals[f"D{r}"] = bal
    vals[f"E{r}"] = E
    vals[f"F{r}"] = F
    vals[f"G{r}"] = G
    vals[f"H{r}"] = fee
    vals[f"I{r}"] = Itot
    vals[f"J{r}"] = J
    vals[f"K{r}"] = Itot
    schedule_rows.append([_add_months(first_due, i), rate, bal,
                          E, F, G, fee, Itot, J])
    cashflows.append(Itot)
    sum_F += F
    sum_G += G
    sum_H += fee
    sum_I += Itot
    bal = J

    # scenario columns L–Q (mirror the spreadsheet formulas)
    M_open = scen_bal
    L_adj = (-WINDOW_AMOUNT if term == start_term else 0.0) + \
            (WINDOW_AMOUNT if term == end_term else 0.0)
    adj = M_open + L_adj
    if adj <= 0:
        N_s = P_s = 0.0
    else:
        N_s = adj * mr * (FIRST_FACTOR if i == 0 else 1.0)
        P_s = max(0.0, adj - (E - N_s))
    O_s = 0.0 if M_open <= 0 else (F - N_s)
    Q_s = J - P_s
    vals[f"L{r}"] = L_adj
    vals[f"M{r}"] = M_open
    vals[f"N{r}"] = N_s
    vals[f"O{r}"] = O_s
    vals[f"P{r}"] = P_s
    vals[f"Q{r}"] = Q_s
    scen_int += N_s
    if P_s > 1:
        scen_payoff_count += 1
    if term == end_term:
        forsprang_after = Q_s
    scen_bal = P_s

vals[f"K{HEAD_ROW}"] = -P
vals["B11"] = _ceil2((1 + rate / tpy) ** tpy - 1)        # rounded up to 2 decimals
# effektiv rente incl. fee via monthly IRR (bisection), annualised
lo, hi = 1e-6, 0.02
for _ in range(200):
    mid = (lo + hi) / 2
    npv = sum(c / (1 + mid) ** k for k, c in enumerate(cashflows))
    lo, hi = (mid, hi) if npv > 0 else (lo, mid)
vals["B12"] = _ceil2((1 + (lo + hi) / 2) ** 12 - 1)      # rounded up to 2 decimals
vals["B13"] = _serial(_add_months(first_due, n - 1))
vals["B14"] = sum_F
vals[f"F{tot}"] = sum_F
vals[f"G{tot}"] = sum_G
vals[f"H{tot}"] = sum_H
vals[f"I{tot}"] = sum_I

# scenario helper + summary cells (D6/D7 dates, D8/D9 term #s, D11–D14 summary)
scen_payoff_term = scen_payoff_count + 1
vals["D6"] = _serial(win_start)
vals["D7"] = _serial(win_end)
vals["D8"] = start_term
vals["D9"] = end_term
vals["D11"] = _serial(_add_months(first_due, scen_payoff_count))   # payoff date
vals["D12"] = n - scen_payoff_term                                # months saved
vals["D13"] = sum_F - scen_int                                    # total interest saved
vals["D14"] = forsprang_after if forsprang_after is not None else 0.0
vals[f"N{tot}"] = scen_int                                        # totals-row scenario interest


def _fmt(v):
    return str(v) if isinstance(v, int) else repr(v)


def _no(v):                       # Norwegian numeric string (comma decimal)
    return f"{v:.2f}".replace(".", ",")


if WANT_XLSX:
    wb.save(XLSX_OUT)
    with zipfile.ZipFile(XLSX_OUT) as zin:
        names = zin.namelist()
        blobs = {nm: zin.read(nm) for nm in names}

    # inject cached values into the model sheet
    xml = blobs["xl/worksheets/sheet1.xml"].decode("utf-8")
    pat = re.compile(r'(<c r="([A-Z]+\d+)"[^>]*>)(<f>[^<]*</f>)<v ?/>')
    counter = [0]

    def _repl(mo):
        coord = mo.group(2)
        if coord in vals:
            counter[0] += 1
            return f"{mo.group(1)}{mo.group(3)}<v>{_fmt(vals[coord])}</v>"
        return mo.group(0)

    blobs["xl/worksheets/sheet1.xml"] = pat.sub(_repl, xml).encode("utf-8")

    # force a full recalc on open (so later edits refresh everything)
    wbxml = blobs["xl/workbook.xml"].decode("utf-8")
    if "fullCalcOnLoad" not in wbxml:
        if "<calcPr" in wbxml:
            wbxml = wbxml.replace("<calcPr ", '<calcPr fullCalcOnLoad="1" ', 1)
        else:
            wbxml = wbxml.replace("</workbook>", '<calcPr fullCalcOnLoad="1"/></workbook>')
        blobs["xl/workbook.xml"] = wbxml.encode("utf-8")

    tmp = XLSX_OUT + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for nm in names:
            zout.writestr(nm, blobs[nm])
    os.replace(tmp, XLSX_OUT)
    print(f"Wrote {XLSX_OUT} ({', '.join(wb.sheetnames)}) · injected {counter[0]} cached values")

if WANT_CSV:
    # flat snapshot of the dynamic monthly schedule (computed values)
    with open(CSV_OUT, "w", encoding="utf-8-sig", newline="") as fh:
        w = csvmod.writer(fh, delimiter=";")
        w.writerow(["Forfallsdato", "Nominell årsrente", "Inngående saldo",
                    "Terminbeløp", "Rente", "Avdrag", "Gebyr", "Totalt", "Restgjeld"])
        for d, rt, D, E, F, G, H, Itot, J in schedule_rows:
            w.writerow([d.strftime("%d.%m.%Y"),
                        f"{rt * 100:.2f}".replace(".", ",") + " %",
                        _no(D), _no(E), _no(F), _no(G), _no(H), _no(Itot), _no(J)])
    print(f"Wrote {CSV_OUT} · {len(schedule_rows)} rows")
