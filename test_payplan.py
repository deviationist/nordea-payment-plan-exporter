"""Integration tests: our computed schedule vs Nordea's published plan.

Rebuilds the workbook, reads the *cached* values from the "Modell (dynamisk)"
sheet (data_only=True — i.e. exactly what a viewer shows), and compares them to
the bank's own `nordea_payplan_data.json`:

  - monthly bank rows (15.07.2026 … 15.12.2027) -> per-term comparison
  - yearly bank rows (31.12.2028 … 31.12.2056) -> compare the SUM of our months
    in that calendar year, plus the year-end restgjeld

Nordea rounds every published row to whole kroner while our model keeps full
precision, so equality is asserted within a small tolerance; the actual max
deviations are printed so you can see how tightly it correlates.

Run:  .venv/bin/python -m unittest -v test_payplan
"""
import json
import subprocess
import sys
import unittest
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
XLSX = HERE / "nordea-nedbetalingsplan.xlsx"
BANK_SRC = HERE / "nordea_payplan_data.json"
SHEET = "Modell (dynamisk)"

# Tolerances (kr). Per-row payment components match to ~1 kr (whole-krone
# rounding). The running balance drifts slowly because the bank's steady payment
# is the whole-krone 29 246 vs our exact annuity 29 245,58 — the bank collects
# ~0,42 kr/month more (all principal), so its balance falls imperceptibly faster
# and the gap grows ~1 kr per elapsed month (≈333 kr after 30 yr, 0,006 %).
TOL_MONTHLY = 1.5        # per-term total/avdrag/rente
FIRST_DUE = date(2026, 7, 15)


def _bal_tol(d):
    """Balance envelope: ~1 kr per elapsed month + small base (payment-rounding drift)."""
    months = (d.year - FIRST_DUE.year) * 12 + (d.month - FIRST_DUE.month)
    return 1.0 * months + 2.0




def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return date(1899, 12, 30) + timedelta(days=int(round(v)))


@unittest.skipUnless((HERE / "nordea_payplan_data.json").exists(),
                     "no downloaded data — run ./fetch.sh first")
class BankComparison(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        subprocess.run([sys.executable, "build_sheet.py", "--format", "xlsx"],
                       cwd=HERE, check=True, capture_output=True)
        ws = load_workbook(XLSX, data_only=True)[SHEET]
        cls.rows = {}
        for r in range(16, 376):
            d = _to_date(ws.cell(r, 2).value)
            cls.rows[d] = dict(
                terminbelop=ws.cell(r, 5).value, rente=ws.cell(r, 6).value,
                avdrag=ws.cell(r, 7).value, gebyr=ws.cell(r, 8).value,
                total=ws.cell(r, 9).value, restgjeld=ws.cell(r, 10).value)
        plans = json.loads(BANK_SRC.read_text())["pay_plans"]
        cls.bank_monthly = [p for p in plans if not p["due_date"].endswith("-12-31")]
        cls.bank_yearly = [p for p in plans if p["due_date"].endswith("-12-31")]

    # ---- monthly ----
    def test_monthly_rows_match_bank(self):
        worst = {}
        for p in self.bank_monthly:
            d = date.fromisoformat(p["due_date"])
            ours = self.rows[d]
            checks = [("total", p["total"], ours["total"], TOL_MONTHLY),
                      ("avdrag", p["amount"], ours["avdrag"], TOL_MONTHLY),
                      ("rente", p["interest"], ours["rente"], TOL_MONTHLY),
                      ("gebyr", p["fee"], ours["gebyr"], 0.01),
                      ("restgjeld", p["loan_balance"], ours["restgjeld"], _bal_tol(d))]
            for name, bank_v, our_v, tol in checks:
                dev = abs(bank_v - our_v)
                worst[name] = max(worst.get(name, 0), dev)
                self.assertLessEqual(dev, tol, f"{d} {name}: bank {bank_v} vs {our_v:.2f}")
        print("\n  monthly max deviation (kr):",
              ", ".join(f"{k}={v:.2f}" for k, v in worst.items()))

    # ---- yearly ----
    def test_yearly_aggregates_match_bank(self):
        worst = {}
        for p in self.bank_yearly:
            y = date.fromisoformat(p["due_date"]).year
            months = [v for d, v in self.rows.items() if d.year == y]
            self.assertTrue(months, f"no model rows in {y}")
            agg = {k: sum(m[k] for m in months) for k in ("total", "avdrag", "rente", "gebyr")}
            year_end_bal = max((d for d in self.rows if d.year == y))
            # every yearly figure must match within the accumulated payment-rounding
            # envelope to date (largest in the payoff year, where the residual lands).
            yt = _bal_tol(year_end_bal)
            checks = [("total", p["total"], agg["total"], yt),
                      ("avdrag", p["amount"], agg["avdrag"], yt),
                      ("rente", p["interest"], agg["rente"], yt),
                      ("gebyr", p["fee"], agg["gebyr"], 1.0),
                      ("restgjeld", p["loan_balance"], self.rows[year_end_bal]["restgjeld"], yt)]
            for name, bank_v, our_v, tol in checks:
                dev = abs(bank_v - our_v)
                worst[name] = max(worst.get(name, 0), dev)
                self.assertLessEqual(dev, tol, f"{y} {name}: bank {bank_v} vs {our_v:.2f}")
        print("\n  yearly max deviation (kr):",
              ", ".join(f"{k}={v:.2f}" for k, v in worst.items()))

    def test_total_interest_correlates(self):
        bank_total_interest = sum(p["interest"] for p in self.bank_monthly + self.bank_yearly)
        our_total_interest = sum(v["rente"] for v in self.rows.values())
        rel = abs(bank_total_interest - our_total_interest) / bank_total_interest
        print(f"\n  total interest: bank {bank_total_interest:,.0f} vs "
              f"ours {our_total_interest:,.0f}  (rel {rel*100:.3f}%)")
        self.assertLess(rel, 0.001)   # within 0.1 %


@unittest.skipUnless((HERE / "nordea_payplan_after_downpayment.json").exists(),
                     "drop nordea_payplan_after_downpayment.json (bank's recalculated "
                     "plan, captured after the downpayment lands) to enable this")
class ScenarioVsBankRecalc(unittest.TestCase):
    """Part 2 — activates once we capture the bank's recalculated plan.

    NOTE: this only correlates if Nordea keeps the monthly payment fixed and
    shortens the term (matching our 'keep paying the original amount' model). If
    the bank instead re-annuitises to a *lower* payment over the same end date,
    the balances will diverge — and that divergence is itself the finding.
    Compare against a *permanent* extra paydown (window with no redraw), not the
    in/out window scenario.
    """
    def test_scenario_balance_tracks_bank_recalc(self):
        self.skipTest("scaffold — implement once the post-payment plan is captured")


if __name__ == "__main__":
    unittest.main(verbosity=2)
